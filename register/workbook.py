from typing import Any, Generator, cast
from openpyxl import load_workbook
from openpyxl.workbook.workbook import Workbook
from openpyxl.worksheet.worksheet import Worksheet

from .helpers import exit_with_info, user_input
from .models import AppSettings


INDEXES = {
    'ordinal_number': 0,
    'financial_source': 3,
    'unit': 12,
    'date': 11,
    'name_of_item': 6,
    'invoice': 4,
    'invoice_date': 5,
    'issuer': 10,
    'value': 8,
    'material_duty_person': 13,
    'psp': 2,
    'cost_center': 2,
    'inventory_number': 2,
    'use_purpose': 14,
    'serial_number': 15,
    'id_vim': 1,
}


def setup_workbook(app_settings: AppSettings, files: list[str]) -> bool:
    """
    Sets up the workbook and sheet name based on the real name.
    Some worksheets have unset columns boundary, we fix that too.

    Parameters:
    app_settings: The AppSettings object.
    files (list[str]): The list of Excel files found in your home directory,
    yes, those on your mounted network filesystems too.

    Returns:
    bool: True if the everything was set successfully, False otherwise.
    """
    index = user_input(
        files,
        'Choose a file which is your workbook you want to use',
    )
    app_settings.wb_filename = files[index]
    workbook = get_workbook(app_settings.wb_filename)
    if workbook:
        if len(workbook.sheetnames) > 1:
            index = user_input(
                workbook.sheetnames,
                'Choose a sheet you want to use',
            )
            app_settings.sheetname = workbook.sheetnames[index]
        app_settings.last_column = obtain_last_data_column_from_worksheet(
            workbook['Środki Trwałe']
        )
        return True
    return False

def get_workbook(filename) -> Workbook:
    """
    Loads an Excel file from the given location on your filesystem.

    Returns:
    Workbook: The current Excel file, or stops if it is not found.
    """

    try:
        workbook = load_workbook(filename, read_only=True)
    except FileNotFoundError:
        exit_with_info(
            f'Cannot find {filename}.\nPlease check your settings.'
        )
    return workbook

def obtain_cell_values_from_worksheet(
        sheet: Worksheet, max_col) -> list[dict] | None:
    """
    Returns cell values from given dimensions

    Parameters:
    sheet (Worksheet): The sheet we retrieve data from.
    max_col (int): The last column in the given sheet obtained from
    the function 'obtain_last_data_column_from_worksheet'.
    """

    rows = sheet.iter_rows(2, max_col=max_col, values_only=True)
    return list(process_rows(rows))

def process_rows(
        rows: Generator[tuple[Any], None, None]
    ) -> Generator[dict, None, None]:
    """
    Processes excel's rows in order defined in INDEXES
    Returns a generator of dictionaries.
    """
    for row in rows:
        yield {key: row[index] for key, index in INDEXES.items()}

def obtain_last_data_column_from_worksheet(sheet: Worksheet) -> int:
    """
    Parameters:
    sheet (Worksheet): The current sheet of the Excel file.

    Returns:
    int: The last column in the given sheet

    Notice that this function is counting from 1 as openpyxl does.
    """

    if row := next(sheet.iter_rows(1, 1, 1, values_only=True), None):
        return next(
            (index - 1 for index, col in enumerate(row, 1) if col is None),
            len(row)
        )
    return 0

def read_workbook_data() -> list[dict[Any]]:
    """
    Reads the data from the workbook and returns it as a list of dictionaries.
    This is an aesy way to import data from a new workbook - simply remove
    the 'wb_filename' entry from setting.txt dictionary stored on your disk
    and start the program adding 'wb' as the parameter.
    """
    app_settings = AppSettings()
    if app_settings.wb_filename is None:
        files = app_settings.list_excel_files()
        setup_workbook(app_settings, files)

    workbook: Workbook = get_workbook(app_settings.wb_filename)  # type: ignore
    rows = obtain_cell_values_from_worksheet(
        cast(Worksheet, workbook[app_settings.sheetname]),
        app_settings.last_column
    )
    workbook.close()
    return rows
