from datetime import datetime
from os import _exit
from pickle import dump, load
from re import match
from typing import Any, cast, overload

from openpyxl import load_workbook
from openpyxl.workbook.workbook import Workbook
from openpyxl.worksheet.worksheet import Worksheet

from .models import App_Settings, FixedAsset
from .financial_sources import FINANCIAL_SOURCES


FILE_DB = 'fixed_assets.db'
DATE_PATTERN = r'^\d{2}-\d{2}-\d{4}$'


def exit_with_info(info: str) -> None:
    """
    Prints an error message and exits with code 1
    """
    print(info)
    _exit(1)

@overload
def user_input(entries: list[str], msg: str) -> int | None: ...
@overload
def user_input(entries: str, msg: str | None = None) -> str | None: ...

def user_input(entries: list[str] | str, msg: str) -> int | str | None:
    """
    Returns the user input depending on the input type
    """
    if isinstance(entries, list):
        for index, file in enumerate(entries, 1):
            print(f'[{index}] {file}')
        return int(input(f'{msg}: ')) - 1
    if isinstance(entries, str):
        return input(
            f'{msg}\n{entries}\nPress enter to confirm or choose another filename: ')
    return None

def setup_wordbook(app_settings: App_Settings, files: list[str]) -> bool:
    index = user_input(
        files,
        'Choose a file which is your wordbook you want to use',
    )
    app_settings.wb_filename = files[index]
    wordbook = get_wordbook(app_settings.wb_filename)
    if wordbook:
        app_settings.sheet_name = wordbook.active.title
        print('The sheet name has been set to: ' + app_settings.sheet_name)
        choice = input('Do you want to change it? (y/n): ')
        if choice.lower() == 'y':
            app_settings.sheet_name = input('Enter new sheet name: ')
        app_settings.last_column = obtain_last_column_from_worksheet(wordbook.active)
        return True
    return False

def get_app_settings() -> App_Settings:
    """
    Returns the complete App_Settings object depending on users input.
    """
    app_settings = App_Settings()

    try:
        files = app_settings.list_files()
    except FileNotFoundError:
        exit_with_info('Cannot find any Excel file.')

    if app_settings.wb_filename == '':
        setup_wordbook(app_settings, files)

    if app_settings.fa_filename == '':
        index = user_input(
            files,
            'Choose a file which is your fixed asset template you want to use',
        )
        app_settings.fa_filename = files[index]

    fa_path = user_input(
        app_settings.fa_path,
        'Enter the path where your fixed asset documents will be stored, currently set to: '
    )
    if fa_path != '' and fa_path != app_settings.fa_path:
        app_settings.fa_path = fa_path

    app_settings.configured = True

    return app_settings

def get_wordbook(filename) -> Workbook | None:
    """
    Loads an Excel file from the given location on disk.

    Returns:
    Workbook: The current Excel file, or stops if it is not found.
    """

    try:
        return load_workbook(filename, read_only=True)
    except FileNotFoundError:
        exit_with_info(
            f'Cannot find {filename}.\nPlease check your settings.')

def obtain_cell_values_from_wordbook(wb: Workbook, max_col) -> list:
    """
    Returns cell values from given dimensions

    Parameters:
    wb (Workbook): The current Excel file.
    max_col (int): The last column in the given sheet obtained from
    the function below.
    """

    sheet: Worksheet = cast(Worksheet, wb.active)
    return list(sheet.iter_rows(2, max_col=max_col, values_only=True))

def obtain_last_column_from_worksheet(sheet: Worksheet) -> int:
    """
    Parameters:
    sheet (Worksheet): The current sheet of the Excel file.

    Returns:
    int: The last column in the given sheet

    Notice that this function is counting from 1 as openpyxl does.
    """

    row = next(sheet.iter_rows(1, 1, 1, values_only=True), None)

    if row:  # if row is not None:
        for index, col in enumerate(row, 1):
            if col is None:
                return index
    return 0

def skip_on_pattern(value: str) -> bool:
    """
    True if data belongs to a group that makes an asset - marked as 'do '

    Parameters:
    value (str): The current cell value.

    Returns:
    bool: True if data belongs to a group that makes an asset, False otherwise.
    """
    return value[:3] == 'do '

def create_document_name(row: tuple) -> str | None:
    """
    Unit and six last digits from the inventory number.
    Returns None if serial contains something else than digits, as it means
    uncomplete stuff - something is currently being built and its costs are
    unknown yet.

    Parameters:
    row (tuple): The current row of the Excel file,
    from which we can obtain the inventory number and unit.

    Returns:
    str | None: The document name if it is valid, None otherwise.
    """
    inventory_number = row[2]
    unit = row[12]
    if inventory_number is None:
        return None
    if unit is None:
        exit_with_info(
            f"""Unit cannot be empty!\n
            Please check your data at ordinal number = {row[0]}.""")
    serial = inventory_number[-6:]
    for char in serial:
        if not char.isdigit():
            return None
    return (unit, serial)

def correct_date(date: str) -> str:
    """
    Corrects the date format by removing any comma or space,
    replaces dots with dashes and returns the corrected date.

    Parameters:
    date (str): The input date string to be corrected.

    Returns:
    str: The corrected date string.
    """
    if ',' in date or ' ' in date:
        if ',' in date:
            date, _ = date.split(',')
        else:
            date, _ = date.split(' ')
    if '.' in date:
        date = date.replace('.', '-')
    return date

def fix_date(_date: Any) -> str:
    """
    Usually _date is a datetime object, but sometimes may be
    given as a string not complying with the actual standards
    (e.g. as 'date,date' literals) in this case the former literal is taken.

    Parameters:
    _date (str | None): The input date string to be corrected.

    Returns:
    str: The corrected date string.
    """
    if _date is None:
        _date = ''
    else:
        try:
            _date = _date.strftime('%d-%m-%Y')
        except AttributeError:
            date_string = correct_date(_date.strip())
            _date = match(DATE_PATTERN, date_string)
            if _date is None:
                raise ValueError(f'Invalid date format: {_date}')

            _date = _date.group()
    return _date

def financial_cost_values(financial_source: str) -> list[str, str] | None:
    """
    Returns a list of psp and cost_center elements if financial_source is found
    in the FINANCIAL_SOURCES dictionary. None oterwise.
    """
    if financial_source in FINANCIAL_SOURCES:
        return [
            FINANCIAL_SOURCES[financial_source]['psp'],
            FINANCIAL_SOURCES[financial_source]['cost_center'],
        ]
    return None

def create_fixed_asset(row: tuple[Any]) -> FixedAsset:
    """
    If financial source (row[3]) is given (which is mostly true),
    it is translated to psp and cost_center repectivly

    Parameters:
    row (tuple): The current row of the Excel file.

    Returns:
    FixedAsset: The created FixedAsset object.
    """
    financial_source = row[3]
    if financial_source is None:
        psp = cost_center = ''
    else:
        source = financial_cost_values(financial_source)
        if source is None:
            psp = cost_center = str(financial_source)
        else:
            psp, cost_center = source

    try:
        date = fix_date(row[11])
        invoice_date = fix_date(row[5])
    except ValueError:
        exit_with_info(f'Please check your data for ordinal number: {row[0]}')

    return FixedAsset(
        date=date,
        name_of_item=row[6],
        invoice=str(row[4]),
        invoice_date=invoice_date,
        issuer=str(row[10]),
        value=str(row[8]),
        material_duty_person=str(row[13]),
        psp=psp,
        cost_center=cost_center,
        inventory_number=row[2],
        )

def select_fixed_asset_elements(rows: list[tuple]) -> list[tuple, FixedAsset]:
    """
    The wordbook I was given had following 16 columns:
    #. ordinal number: it represents an invoice or a group of them,
       may repeat itself many times,
       may also be skipped (None), AFTER it has been specified once.
    #. unused here
    #. inventory number
    #. financial source - probably the most important column in thw wordbook,
       as it constitutes the fields of psp and cost_center.
    #. invoice number
    #. invoice date
    #. name of a product/asset
    #. quantity (uusally 1)
    #. price
    #. value of the two above (formula), only this column is used here
    #. producent/supplier
    #. registering date - when the asset was accepted to the register
    #. unit - the devision an asset belongs to
    #. material duty person
    #. 15-16 unused here
    """

    selected_elements = []
    for i, row in enumerate(rows, 1):
        ordinal_number = row[0]
        inventory_number = row[2]

        # is_pattern = skip_on_pattern(inventory_number)
        if (ordinal_number is None or inventory_number is None
            or skip_on_pattern(inventory_number)
            # or is_pattern
            and i > 1 and ordinal_number == rows[i - 1][0]):
            continue

        document_name_tuple = create_document_name(row)
        if document_name_tuple is not None:
            try:
                fixed_asset = create_fixed_asset(row)
            except TypeError:
                # shouldn,t happen but it's a good way to find
                # data which might cause a problem
                fixed_asset = None
            selected_elements.append([document_name_tuple, fixed_asset])
    return selected_elements

def final_data_checking(serial: str) -> None:
    """
    Checks if serial number is empty or None. If something is wrong,
    here, we made an error before.
    """
    if serial is None or serial == '':
        raise ValueError("""Serial number is empty!\n
                            Please check your data for serial number""")

def check_serials(elemets: list[tuple, FixedAsset]) -> list[FixedAsset] | None:
    """
    Returns None if all double_elements are unique, otherwise a list
    of doubled elements.
    """
    serials = {}
    for t, _ in elemets:
        _, serial = t
        final_data_checking(serial)
        serials[serial] = serials.get(serial, 0) + 1
    doubles = [[k, v] for k, v in serials.items() if v > 1]
    if len(doubles) > 0:
        return doubles
    return None

def read_wordbook_data() -> list[tuple]:
    app_settings = App_Settings()
    wordbook: Workbook = get_wordbook(app_settings.wb_filename)  # type: ignore
    rows = obtain_cell_values_from_wordbook(wordbook, app_settings.last_column)
    wordbook.close()
    return rows

def print_double_elements(double_elements: list[list], selected_items: list) -> None:
    print('The following serial numbers are repeated this number of times:')
    for d in double_elements:
        print(f'{d[0]}: {d[1]}')
        for t, fa in selected_items:
            u, s = t
            if s == d[0]:
                print(f'\t{u}\n{fa}')
    print('Please check your data and try again.')

def process_workbook_data(rows: list[tuple]) -> None:
    """
    Imports selected data from a wordbook and stores it
    in a pickle DB file if there is no doubled elements (serials).
    The latter means error in the provided data, so no dump is done.
    """
    selected_items = select_fixed_asset_elements(rows)
    double_elements = check_serials(selected_items)
    if double_elements:
        print_double_elements(double_elements, selected_items)
    else:
        with open(FILE_DB, 'wb') as stream:
            dump(selected_items, stream)

def load_fixed_assets() -> list[FixedAsset]:
    try:
        with open(FILE_DB, 'rb') as reader:
            fixed_assets = load(reader, encoding='utf-8')
    except FileNotFoundError:
        exit_with_info(f'File \'{FILE_DB}\' cannot be opened.')
    if fixed_assets is None:
        return []
    else:
        return fixed_assets

def print_fixed_assets(
        fixed_assets_elements: list[str, tuple[str], FixedAsset]
        ) -> None:
    print()
    # for ordinal, doc_name, fixed_asset in fixed_assets_elements:
    for doc_name, ordinal, fixed_asset in fixed_assets_elements:  
        unit, serial = doc_name
        print(f'{ordinal}, {unit}-{serial}\n', fixed_asset)
